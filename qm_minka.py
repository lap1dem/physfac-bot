# -*- coding: utf-8 -*-

# %%

import os
import random

import numpy as np
from openpyxl import load_workbook
ex = os.path.abspath("Мінка кванти.xlsx")
wb = load_workbook(ex)
cons = wb[wb.sheetnames[0]]


# %%
def get_question(nsem):
    if nsem == '1 семестр':
        sem = list(np.arange(3,74))
    elif nsem == '2 семестр':
        sem = list(np.arange(75,138))
    else:
        sem = list(np.arange(3,74))+list(np.arange(75,138))
    q = random.choice(sem)
    question =  cons['A'+ str(q)].value + ". " + cons['B'+ str(q)].value
    return(question)
