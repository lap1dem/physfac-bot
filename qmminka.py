# -*- coding: utf-8 -*-

# %%

import os
import random

import numpy as np
from openpyxl import load_workbook
cp = os.getcwd()
wb = load_workbook(filename = cp+'/Мінка кванти.xlsx')
cons = wb[wb.sheetnames[0]]

# %%
def get_question():
    sem = list(np.arange(3,74))+list(np.arange(75,138))
    q = random.choice(sem)
    question =  cons['A'+ str(q)].value + ". " + cons['B'+ str(q)].value
    return(question)
